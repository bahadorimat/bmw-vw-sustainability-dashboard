"""
German Automotive Sustainability Intelligence Dashboard
BMW Group vs. Volkswagen Group — CSRD & KPI Analysis (FY 2025)

Matin Bahadori, 2025
M.Sc. Umwelttechnik — BTU Cottbus-Senftenberg

Data sources:
  BMW Group Report 2025 (p. 128–129, energy/water/waste sections)
  Volkswagen Group ESRS Sustainability Report 2025 (p. 273, KPI tables)

Standard: EU CSRD / ESRS Set 1 (2023)
Auditors: PricewaterhouseCoopers GmbH (BMW), Ernst & Young GmbH (VW)
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
from matplotlib.patches import FancyBboxPatch
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = "output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Colors ────────────────────────────────────────────────────────────────────
BMW_BLUE = "#1C69D4"
VW_TEAL  = "#00B0F0"
BMW_DARK = "#003399"
VW_DARK  = "#006EA6"
DARK_TEXT = "#1A1A1A"
MID_TEXT  = "#666666"
GRID_COL  = "#E8E8E8"

plt.rcParams.update({
    "font.family":       "DejaVu Sans",
    "font.size":         10,
    "axes.spines.top":   False,
    "axes.spines.right": False,
    "axes.edgecolor":    "#444444",
    "axes.linewidth":    0.8,
})

# ════════════════════════════════════════════════════════════════════════════
# REAL DATA — extracted from official sustainability reports
# ════════════════════════════════════════════════════════════════════════════

# ── GHG Emissions (Mt CO2e) ───────────────────────────────────────────────────
# BMW: BMW Group Report 2025, p. 128–129
# VW:  VW ESRS Sustainability Report 2025, p. 273 (Scope 1+2 market-based)

ghg = pd.DataFrame({
    "metric":   ["Scope 1+2\n(market-based)", "Scope 3\nTotal", "Total\nAll Scopes"],
    "BMW_2025": [0.811,  127.54, 128.35],
    "BMW_2024": [0.837,  134.98, 135.82],
    "VW_2025":  [2.80,   883.74, 886.54],
    "VW_2024":  [3.50,   824.00, 827.50],
})

# ── Scope 3 Category Breakdown (Mt CO2e) ─────────────────────────────────────
# BMW: BMW Group Report 2025, p.128–129
# VW:  VW ESRS Sustainability Report 2025, p. 273

scope3_cats = pd.DataFrame({
    "category": [
        "Cat 1 — Purchased\nGoods & Services",
        "Cat 2 — Capital\nGoods",
        "Cat 3–5 — Fuel,\nTransport & Waste",
        "Cat 6–8 — Travel,\nCommuting & Leased",
        "Cat 11 — Use of\nSold Products",
        "Cat 12 — End-of-\nLife Treatment",
        "Cat 13–15 — Leased\nAssets, Franchises",
    ],
    "BMW_2025": [31.71,  1.625, 2.451+0.0,  0.063+0.172,     89.539, 1.425, 0.556],
    "VW_2025":  [104.44, 8.33,  1.29+6.73+1.64, 0.22+0.31+0.18, 752.57, 0.91, 4.42+2.49+0.01],
})

# ── KPI Scorecard ─────────────────────────────────────────────────────────────
# BMW: BMW Group Report 2025 (energy, water, waste sections)
# VW:  VW ESRS Sustainability Report 2025 (KPI tables)

kpis = {
    "Total Energy (million MWh)": {
        "BMW_2025": 6.18, "BMW_2024": 6.21,
        "VW_2025": 19.2,  "VW_2024": 19.0,
        "unit": "million MWh", "lower_is_better": True,
    },
    "Renewable Energy Share (%)": {
        "BMW_2025": 49.4, "BMW_2024": 48.5,
        "VW_2025": 40.6,  "VW_2024": 37.4,
        "unit": "%", "lower_is_better": False,
    },
    "Water Withdrawal (million m³)": {
        "BMW_2025": 5.70, "BMW_2024": 5.85,
        "VW_2025": 19.9,  "VW_2024": 21.2,
        "unit": "million m³", "lower_is_better": True,
    },
    "Total Waste (thousand tonnes)": {
        "BMW_2025": 851.8,  "BMW_2024": 873.4,
        "VW_2025": 2547.3,  "VW_2024": 2357.7,
        "unit": "thousand t", "lower_is_better": True,
    },
    "Employees (thousands)": {
        "BMW_2025": 154.5, "BMW_2024": 157.5,
        "VW_2025": 602.7,  "VW_2024": 614.1,
        "unit": "thousands", "lower_is_better": False,
    },
    "Scope 3 per Employee (kt CO2e/emp)": {
        "BMW_2025": round(127.54/154.5*1000, 1),
        "BMW_2024": round(134.98/157.5*1000, 1),
        "VW_2025":  round(883.74/602.7*1000, 1),
        "VW_2024":  round(824.00/614.1*1000, 1),
        "unit": "t CO2e/employee", "lower_is_better": True,
    },
}

# ── CSRD / ESRS Reporting Status ──────────────────────────────────────────────
esrs_topics = pd.DataFrame({
    "topic": [
        "E1 — Climate Change",
        "E2 — Pollution",
        "E3 — Water & Marine Resources",
        "E4 — Biodiversity & Ecosystems",
        "E5 — Resource Use & Circular Economy",
        "S1 — Own Workforce",
        "S2 — Value Chain Workers",
        "S3 — Affected Communities",
        "S4 — Consumers & End-Users",
        "G1 — Business Conduct",
    ],
    "pillar": ["E","E","E","E","E","S","S","S","S","G"],
    "BMW_reports":   [1, 1, 1, 1, 1, 1, 1, 0, 0, 1],
    "BMW_material":  [1, 1, 0, 0, 1, 1, 1, 0, 0, 1],
    "VW_reports":    [1, 1, 1, 1, 1, 1, 1, 1, 1, 1],
    "VW_material":   [1, 1, 1, 1, 1, 1, 1, 1, 0, 1],
})

print("Data loaded successfully.")
print(f"\nBMW Total Scope 3 (2025): {ghg.loc[1,'BMW_2025']:.2f} Mt CO2e")
print(f"VW  Total Scope 3 (2025): {ghg.loc[1,'VW_2025']:.2f} Mt CO2e")
print(f"\nVW Scope 3 is {ghg.loc[1,'VW_2025']/ghg.loc[1,'BMW_2025']:.1f}x larger than BMW")


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 1 — GHG Emissions Comparison
# ════════════════════════════════════════════════════════════════════════════

fig1, axes = plt.subplots(1, 3, figsize=(18, 7))
fig1.patch.set_facecolor("white")

fig1.suptitle(
    "Figure 1.  GHG Emissions Comparison — BMW Group vs. Volkswagen Group (FY 2025)",
    fontsize=13, fontweight="bold", y=0.98, color=DARK_TEXT)
fig1.text(0.5, 0.93,
    "Data: BMW Group Report 2025 (p.128–129) | VW ESRS Sustainability Report 2025 (p.273) | Unit: Mt CO₂e",
    ha="center", fontsize=9, color=MID_TEXT, style="italic")

titles = ["Scope 1+2 (market-based)", "Scope 3 Total", "Total All Scopes"]
for i, (ax, title, row) in enumerate(zip(axes, titles, range(3))):
    ax.set_facecolor("#F8F8F8")
    companies = ["BMW\n2024", "BMW\n2025", "VW\n2024", "VW\n2025"]
    values    = [ghg.loc[row,"BMW_2024"], ghg.loc[row,"BMW_2025"],
                 ghg.loc[row,"VW_2024"],  ghg.loc[row,"VW_2025"]]
    colors    = [BMW_DARK, BMW_BLUE, VW_DARK, VW_TEAL]
    bars = ax.bar(companies, values, color=colors,
                  edgecolor="white", linewidth=1, width=0.55)
    for bar, val in zip(bars, values):
        ax.text(bar.get_x() + bar.get_width()/2, val * 1.02,
                f"{val:.1f}", ha="center", va="bottom",
                fontsize=9, fontweight="bold", color=DARK_TEXT)
    ax.set_title(title, fontsize=11, fontweight="bold", pad=10, color=DARK_TEXT)
    ax.set_ylabel("Mt CO₂e", fontsize=9, color=DARK_TEXT)
    ax.yaxis.grid(True, color=GRID_COL, linewidth=0.7, zorder=0)
    ax.set_axisbelow(True)

    # change arrow
    bmw_chg = (ghg.loc[row,"BMW_2025"] - ghg.loc[row,"BMW_2024"]) / ghg.loc[row,"BMW_2024"] * 100
    vw_chg  = (ghg.loc[row,"VW_2025"]  - ghg.loc[row,"VW_2024"])  / ghg.loc[row,"VW_2024"]  * 100
    ax.text(0.5, -0.13, f"BMW: {bmw_chg:+.1f}%",
            ha="center", transform=ax.transAxes,
            fontsize=9, color=BMW_BLUE, fontweight="bold")
    ax.text(0.5, -0.19, f"VW: {vw_chg:+.1f}%",
            ha="center", transform=ax.transAxes,
            fontsize=9, color=VW_TEAL, fontweight="bold")

fig1.text(0.01, 0.01,
    "Note: Scope 1+2 uses market-based Scope 2 method. "
    "VW Scope 3 2025 increase driven by inclusion of Power Engineering segment (MAN, Traton, Rolls-Royce Power Systems).",
    fontsize=7.5, color=MID_TEXT, style="italic")

plt.tight_layout(rect=[0, 0.06, 1, 0.92])
plt.savefig(f"{OUTPUT_DIR}/fig1_ghg_comparison.png",
            dpi=150, bbox_inches="tight", facecolor="white")
plt.show()
plt.close()
print("Figure 1 saved.")


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 2 — Scope 3 Category Breakdown
# ════════════════════════════════════════════════════════════════════════════

fig2, (ax1, ax2) = plt.subplots(1, 2, figsize=(16, 8))
fig2.patch.set_facecolor("white")

fig2.suptitle(
    "Figure 2.  Scope 3 Category Breakdown — BMW Group vs. Volkswagen Group (FY 2025)",
    fontsize=13, fontweight="bold", y=0.98, color=DARK_TEXT)

cat_colors = ["#E53935","#FF7043","#FDD835","#43A047",
              "#FB8C00","#00ACC1","#546E7A"]

for ax, company, col, title in zip(
    [ax1, ax2],
    ["BMW_2025", "VW_2025"],
    [BMW_BLUE, VW_TEAL],
    ["BMW Group — Scope 3 by Category (2025)",
     "Volkswagen Group — Scope 3 by Category (2025)"]
):
    ax.set_facecolor("white")
    total = scope3_cats[company].sum()
    vals  = scope3_cats[company].values
    cats  = scope3_cats["category"].values

    wedges, texts, autotexts = ax.pie(
        vals, colors=cat_colors,
        autopct=lambda p: f"{p:.1f}%" if p > 1.5 else "",
        startangle=140, pctdistance=0.78,
        wedgeprops={"edgecolor":"white","linewidth":2,"width":0.65},
    )
    for at in autotexts:
        at.set_color("white"); at.set_fontsize(8.5); at.set_fontweight("bold")

    ax.text(0, 0.1, f"{total:.1f}", ha="center",
            fontsize=14, fontweight="bold", color=DARK_TEXT)
    ax.text(0, -0.18, "Mt CO₂e", ha="center",
            fontsize=9, color=MID_TEXT)

    patches = [mpatches.Patch(color=c, label=f"{cat.replace(chr(10),' ')} — {val:.2f} Mt")
               for c, cat, val in zip(cat_colors, cats, vals)]
    ax.legend(handles=patches, loc="lower center",
              bbox_to_anchor=(0.5, -0.45), frameon=False,
              fontsize=7.5, ncol=1)
    ax.set_title(title, fontsize=11, fontweight="bold",
                 pad=12, color=DARK_TEXT)

fig2.text(0.5, 0.01,
    "Source: BMW Group Report 2025 p.128–129 | VW ESRS Sustainability Report 2025 p.273 | "
    "GHG Protocol Corporate Value Chain (Scope 3) Standard",
    ha="center", fontsize=8, color=MID_TEXT, style="italic")

plt.tight_layout(rect=[0, 0.06, 1, 0.95])
plt.savefig(f"{OUTPUT_DIR}/fig2_scope3_breakdown.png",
            dpi=150, bbox_inches="tight", facecolor="white")
plt.show()
plt.close()
print("Figure 2 saved.")


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 3 — KPI Scorecard Comparison
# ════════════════════════════════════════════════════════════════════════════

fig3, axes3 = plt.subplots(2, 3, figsize=(18, 10))
fig3.patch.set_facecolor("white")

fig3.suptitle(
    "Figure 3.  Sustainability KPI Scorecard — BMW Group vs. Volkswagen Group (FY 2024–2025)",
    fontsize=13, fontweight="bold", y=0.99, color=DARK_TEXT)

kpi_items = list(kpis.items())
for idx, (ax, (kpi_name, kpi_vals)) in enumerate(
    zip(axes3.flat, kpi_items)
):
    ax.set_facecolor("#F8F8F8")
    companies = ["BMW\n2024","BMW\n2025","VW\n2024","VW\n2025"]
    values    = [kpi_vals["BMW_2024"], kpi_vals["BMW_2025"],
                 kpi_vals["VW_2024"],  kpi_vals["VW_2025"]]
    colors    = [BMW_DARK, BMW_BLUE, VW_DARK, VW_TEAL]

    bars = ax.bar(companies, values, color=colors,
                  edgecolor="white", linewidth=0.8, width=0.55)

    for bar, val in zip(bars, values):
        label = f"{val:.1f}" if val >= 10 else f"{val:.2f}"
        ax.text(bar.get_x()+bar.get_width()/2, val*1.02,
                label, ha="center", va="bottom",
                fontsize=8.5, fontweight="bold", color=DARK_TEXT)

    ax.set_title(kpi_name, fontsize=10, fontweight="bold",
                 pad=8, color=DARK_TEXT)
    ax.set_ylabel(kpi_vals["unit"], fontsize=8.5, color=DARK_TEXT)
    ax.yaxis.grid(True, color=GRID_COL, linewidth=0.7, zorder=0)
    ax.set_axisbelow(True)

    # YoY change labels
    bmw_chg = (kpi_vals["BMW_2025"]-kpi_vals["BMW_2024"])/kpi_vals["BMW_2024"]*100
    vw_chg  = (kpi_vals["VW_2025"] -kpi_vals["VW_2024"]) /kpi_vals["VW_2024"] *100
    ax.text(0.27, -0.15, f"BMW YoY: {bmw_chg:+.1f}%",
            ha="center", transform=ax.transAxes,
            fontsize=8, color=BMW_BLUE, fontweight="bold")
    ax.text(0.73, -0.15, f"VW YoY: {vw_chg:+.1f}%",
            ha="center", transform=ax.transAxes,
            fontsize=8, color=VW_TEAL, fontweight="bold")

fig3.text(0.01, 0.01,
    "Source: BMW Group Report 2025 | VW ESRS Sustainability Report 2025 | "
    "Renewable % = renewable MWh / total MWh consumed",
    fontsize=7.5, color=MID_TEXT, style="italic")

plt.tight_layout(rect=[0, 0.04, 1, 0.96])
plt.savefig(f"{OUTPUT_DIR}/fig3_kpi_scorecard.png",
            dpi=150, bbox_inches="tight", facecolor="white")
plt.show()
plt.close()
print("Figure 3 saved.")


# ════════════════════════════════════════════════════════════════════════════
# FIGURE 4 — CSRD / ESRS Reporting Comparison
# ════════════════════════════════════════════════════════════════════════════

fig4, ax4 = plt.subplots(figsize=(14, 8))
fig4.patch.set_facecolor("white")
ax4.set_facecolor("white")
ax4.axis("off")

ax4.set_title(
    "Figure 4.  CSRD / ESRS Topic Reporting & Materiality — BMW Group vs. Volkswagen Group (FY 2025)",
    fontsize=12, fontweight="bold", pad=14, color=DARK_TEXT)

cols    = ["ESRS Topic", "Pillar",
           "BMW\nReports", "BMW\nMaterial",
           "VW\nReports",  "VW\nMaterial"]
col_x   = [0.01, 0.38, 0.50, 0.62, 0.74, 0.86]
col_w   = [0.36, 0.10, 0.10, 0.10, 0.10, 0.10]
row_h   = 0.072
start_y = 0.92

pillar_colors = {"E":"#1B5E20","S":"#B71C1C","G":"#1A237E"}

# header
for c, (col, cx) in enumerate(zip(cols, col_x)):
    rect = FancyBboxPatch((cx, start_y-0.005), col_w[c], row_h,
                          boxstyle="round,pad=0.005",
                          facecolor="#1E3A5F", edgecolor="white",
                          linewidth=0.5, transform=ax4.transAxes)
    ax4.add_patch(rect)
    ax4.text(cx + col_w[c]/2, start_y + row_h/2 - 0.005, col,
             ha="center", va="center", fontsize=9,
             fontweight="bold", color="white",
             transform=ax4.transAxes)

for r, row in esrs_topics.iterrows():
    y = start_y - (r+1) * row_h
    bg = "#F5F5F5" if r % 2 == 0 else "white"
    rect = FancyBboxPatch((col_x[0], y), sum(col_w), row_h-0.003,
                          boxstyle="round,pad=0.003",
                          facecolor=bg, edgecolor="#EEEEEE",
                          linewidth=0.3, transform=ax4.transAxes)
    ax4.add_patch(rect)

    # topic name
    ax4.text(col_x[0]+0.01, y+row_h/2, row["topic"],
             va="center", fontsize=9, color=DARK_TEXT,
             fontweight="bold", transform=ax4.transAxes)

    # pillar badge
    pc = pillar_colors.get(row["pillar"], "#555")
    prect = FancyBboxPatch((col_x[1]+0.005, y+0.01),
                           0.05, row_h-0.025,
                           boxstyle="round,pad=0.003",
                           facecolor=pc, edgecolor="none",
                           transform=ax4.transAxes)
    ax4.add_patch(prect)
    ax4.text(col_x[1]+0.03, y+row_h/2, row["pillar"],
             ha="center", va="center", fontsize=9,
             fontweight="bold", color="white",
             transform=ax4.transAxes)

    # checkmarks
    checks = [
        (col_x[2], row["BMW_reports"],  BMW_BLUE),
        (col_x[3], row["BMW_material"], BMW_BLUE),
        (col_x[4], row["VW_reports"],   VW_TEAL),
        (col_x[5], row["VW_material"],  VW_TEAL),
    ]
    for cx, val, color in checks:
        symbol = "●" if val == 1 else "○"
        fc     = color if val == 1 else "#CCCCCC"
        ax4.text(cx + col_w[2]/2, y+row_h/2, symbol,
                 ha="center", va="center", fontsize=13,
                 color=fc, transform=ax4.transAxes)

# legend
legend_y = start_y - (len(esrs_topics)+1.5) * row_h
ax4.text(0.01, legend_y,
    "●  = Yes   ○  = No   |   "
    "Reports = Topic covered in sustainability report   |   "
    "Material = Identified as material under Double Materiality Assessment",
    fontsize=8, color=MID_TEXT, transform=ax4.transAxes)

# summary counts
bmw_rep = esrs_topics["BMW_reports"].sum()
bmw_mat = esrs_topics["BMW_material"].sum()
vw_rep  = esrs_topics["VW_reports"].sum()
vw_mat  = esrs_topics["VW_material"].sum()

ax4.text(0.01, legend_y - 0.07,
    f"BMW Group:  {bmw_rep}/10 topics reported  |  {bmw_mat}/10 material     "
    f"Volkswagen Group:  {vw_rep}/10 topics reported  |  {vw_mat}/10 material",
    fontsize=9, color=DARK_TEXT, fontweight="bold",
    transform=ax4.transAxes)

plt.savefig(f"{OUTPUT_DIR}/fig4_csrd_comparison.png",
            dpi=150, bbox_inches="tight", facecolor="white")
plt.show()
plt.close()
print("Figure 4 saved.")


# ════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT — 3 sheets
# ════════════════════════════════════════════════════════════════════════════

wb  = Workbook()
ws1 = wb.active
ws1.title = "GHG Emissions"
ws1.sheet_view.showGridLines = False

hdr_fill  = PatternFill("solid", start_color="1E3A5F")
bmw_fill  = PatternFill("solid", start_color="E3F2FD")
vw_fill   = PatternFill("solid", start_color="E0F7FA")
bold_wht  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
bold_blk  = Font(name="Arial", bold=True, color="1A1A1A", size=10)
norm_blk  = Font(name="Arial", color="1A1A1A", size=10)
center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
thin      = Side(style="thin", color="CCCCCC")
bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

# GHG sheet
headers = ["Metric","Unit",
           "BMW 2025","BMW 2024","BMW Change",
           "VW 2025","VW 2024","VW Change",
           "Ratio VW/BMW (2025)"]
widths  = [28,14,14,14,14,14,14,14,20]
for c,(h,w) in enumerate(zip(headers,widths),1):
    cell=ws1.cell(1,c,h)
    cell.font=bold_wht; cell.fill=hdr_fill
    cell.alignment=center; cell.border=bdr
    ws1.column_dimensions[get_column_letter(c)].width=w
ws1.row_dimensions[1].height=30

ghg_export = [
    ("Scope 1 (Mt CO2e)", "Mt CO2e", 0.5, 0.55, 2.4, 3.0),
    ("Scope 2 market-based (Mt CO2e)", "Mt CO2e", 0.311, 0.287, 0.4, 0.5),
    ("Scope 1+2 market-based (Mt CO2e)", "Mt CO2e", 0.811, 0.837, 2.80, 3.50),
    ("Scope 3 Total (Mt CO2e)", "Mt CO2e", 127.54, 134.98, 883.74, 824.00),
    ("Total All Scopes (Mt CO2e)", "Mt CO2e", 128.35, 135.82, 886.54, 827.50),
]
for r,(name,unit,b25,b24,v25,v24) in enumerate(ghg_export,2):
    data=[name,unit,b25,b24,f"=D{r}-C{r}",v25,v24,f"=G{r}-F{r}",f"=F{r}/C{r}"]
    fmts=[None,None,"#,##0.00","#,##0.00","+#,##0.00;-#,##0.00",
          "#,##0.00","#,##0.00","+#,##0.00;-#,##0.00","0.0x"]
    for c,(v,fmt) in enumerate(zip(data,fmts),1):
        cell=ws1.cell(r,c,v)
        cell.font=norm_blk
        cell.fill=bmw_fill if c in [3,4,5] else (vw_fill if c in [6,7,8] else PatternFill())
        cell.alignment=center if c!=1 else left
        cell.border=bdr
        if fmt: cell.number_format=fmt
    ws1.row_dimensions[r].height=20

# KPI sheet
ws2 = wb.create_sheet("KPI Scorecard")
ws2.sheet_view.showGridLines = False
kpi_headers = ["KPI","Unit","BMW 2025","BMW 2024","BMW YoY %",
               "VW 2025","VW 2024","VW YoY %","Better performer (2025)"]
kpi_widths  = [35,20,14,14,13,14,14,13,22]
for c,(h,w) in enumerate(zip(kpi_headers,kpi_widths),1):
    cell=ws2.cell(1,c,h)
    cell.font=bold_wht; cell.fill=hdr_fill
    cell.alignment=center; cell.border=bdr
    ws2.column_dimensions[get_column_letter(c)].width=w
ws2.row_dimensions[1].height=30

for r,(kpi_name,kpi_vals) in enumerate(kpis.items(),2):
    b25=kpi_vals["BMW_2025"]; b24=kpi_vals["BMW_2024"]
    v25=kpi_vals["VW_2025"];  v24=kpi_vals["VW_2024"]
    lib=kpi_vals["lower_is_better"]
    better = "BMW" if (b25<v25)==lib else "VW"
    data=[kpi_name, kpi_vals["unit"], b25, b24,
          f"=D{r}/C{r}-1", v25, v24, f"=G{r}/F{r}-1", better]
    fmts=[None,None,"#,##0.00","#,##0.00","0.0%",
          "#,##0.00","#,##0.00","0.0%",None]
    for c,(v,fmt) in enumerate(zip(data,fmts),1):
        cell=ws2.cell(r,c,v)
        cell.font=norm_blk
        cell.fill=bmw_fill if c in [3,4,5] else (vw_fill if c in [6,7,8] else PatternFill())
        cell.alignment=center if c not in [1,9] else left
        cell.border=bdr
        if fmt: cell.number_format=fmt
    ws2.row_dimensions[r].height=22

# Methodology sheet
ws3 = wb.create_sheet("Data Sources")
ws3.sheet_view.showGridLines = False
ws3.column_dimensions["A"].width=30
ws3.column_dimensions["B"].width=45
ws3.column_dimensions["C"].width=45

meta_hdr=["Field","BMW Group","Volkswagen Group"]
for c,h in enumerate(meta_hdr,1):
    cell=ws3.cell(1,c,h)
    cell.font=bold_wht; cell.fill=hdr_fill
    cell.alignment=center; cell.border=bdr

meta=[
    ("Report name","BMW Group Report 2025","VW ESRS Sustainability Report 2025"),
    ("Page reference","p. 128–129 (GHG), energy/water/waste sections","p. 273 (Scope 3), KPI tables"),
    ("Reporting year","FY 2025 (Jan–Dec 2025)","FY 2025 (Jan–Dec 2025)"),
    ("GHG standard","GHG Protocol Scope 3 Standard","GHG Protocol Scope 3 Standard"),
    ("CSRD standard","ESRS Set 1 (2023)","ESRS Set 1 (2023)"),
    ("Auditor","PricewaterhouseCoopers GmbH","Ernst & Young GmbH"),
    ("Audit level","Reasonable assurance","Reasonable assurance"),
    ("Scope 2 method","Market-based","Market-based"),
    ("Vehicles sold (~)","~2.5 million","~9.0 million"),
    ("Employees (2025)","154,540","602,659"),
    ("URL","bmwgroup.com/en/sustainability","volkswagen-group.com/sustainability"),
    ("Analysis by","Matin Bahadori — BTU Cottbus-Senftenberg",""),
]
for r,(a,b,c_val) in enumerate(meta,2):
    for col,val in enumerate([a,b,c_val],1):
        cell=ws3.cell(r,col,val)
        cell.font=bold_blk if col==1 else norm_blk
        cell.fill=PatternFill() if col==1 else (bmw_fill if col==2 else vw_fill)
        cell.alignment=left; cell.border=bdr
    ws3.row_dimensions[r].height=20

path_xl=f"{OUTPUT_DIR}/bmw_vw_sustainability_2025.xlsx"
wb.save(path_xl)
print(f"Excel saved → {path_xl}")


# ── Summary print ──────────────────────────────────────────────────────────
print(f"\n{'='*65}")
print(f"  BMW Group vs. Volkswagen Group — Sustainability Summary 2025")
print(f"{'='*65}")
print(f"  {'Metric':<35} {'BMW':>10}  {'VW':>10}  {'Ratio':>8}")
print(f"  {'-'*65}")
rows=[
    ("Scope 3 (Mt CO2e)",        127.54,  883.74),
    ("Total energy (M MWh)",       6.18,   19.2),
    ("Renewable share (%)",        49.4,   40.6),
    ("Water withdrawal (M m³)",    5.70,   19.9),
    ("Waste (thousand t)",        851.8, 2547.3),
    ("Employees (thousands)",     154.5,  602.7),
    ("Scope3/employee (t/emp)",
     round(127.54/154.5*1e6/1e3,1),
     round(883.74/602.7*1e6/1e3,1)),
]
for name,b,v in rows:
    ratio=v/b
    print(f"  {name:<35} {b:>10.1f}  {v:>10.1f}  {ratio:>7.1f}x")
print(f"{'='*65}")
print(f"\nAll outputs saved to output/ folder.")
